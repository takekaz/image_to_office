
import os
import json
import tempfile
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

def insert_images_to_excel(excel_filepath: str, image_folder_path: str, regions_and_coords: list):
    """
    指定された画像フォルダ内の画像を読み込み、その領域をExcelシートの指定セルに貼り付けます。
    画像ごとに新しいシートを作成します。

    Args:
        excel_filepath (str): 出力するExcelファイルのパス。
        image_folder_path (str): 画像が保存されているフォルダのパス。
        regions_and_coords (list): 領域とセル座標のペアのリスト。
                                   例: [{"img_region": [x1, y1, x2, y2], "excel_pos": "B2"}, ...]
                                   img_region: [left, upper, right, lower] (Pillowのcrop形式)
    """
    # 既存のファイルがあっても上書きで新規作成
    wb = Workbook()

    # デフォルトで作成されるシートを削除（または名前を変更して利用）
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    temp_files_to_delete = [] # 一時ファイルを格納するリスト

    supported_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff')

    # 画像フォルダ内の全ての画像ファイルを処理
    for image_filename in sorted(os.listdir(image_folder_path)):
        if image_filename.lower().endswith(supported_extensions):
            image_path = os.path.join(image_folder_path, image_filename)

            # 画像ごとに新しいシートを作成
            sheet_name = os.path.splitext(image_filename)[0][:31] # シート名は31文字まで
            ws = wb.create_sheet(title=sheet_name)
            print(f"Processing image: {image_filename} on sheet: {sheet_name}")

            try:
                original_image = Image.open(image_path)

                for item in regions_and_coords:
                    img_region = item["img_region"]
                    excel_pos = item["excel_pos"]

                    # 画像領域を切り抜き
                    cropped_image = original_image.crop(img_region)

                    # 切り抜いた画像を一時ファイルとして保存
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
                        temp_image_path = temp_file.name
                        cropped_image.save(temp_image_path)
                        temp_files_to_delete.append(temp_image_path) # リストに追加

                    # ExcelImageオブジェクトを作成し、シートにアンカーして貼り付け
                    img = ExcelImage(temp_image_path)
                    ws.add_image(img, excel_pos)
                    print(f"  - Cropped region {img_region} from {image_filename} and inserted at {excel_pos}")

            except Exception as e:
                print(f"Error processing {image_filename}: {e}")
                continue

    # ワークブックを保存
    try:
        wb.save(excel_filepath)
        print(f"Excel file saved successfully to {excel_filepath}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
    finally:
        # Excelファイルの保存が完了した後、一時ファイルをクリーンアップ
        for temp_file in temp_files_to_delete:
            try:
                os.remove(temp_file)
            except Exception as e:
                print(f"Error deleting temporary file {temp_file}: {e}")

if __name__ == '__main__':
    current_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(current_dir, "config.json")
    image_dir = os.path.join(current_dir, "img")
    output_excel_file = os.path.join(current_dir, "output_images.xlsx")

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)

        regions_and_coords = config.get("image_regions_and_excel_coords", [])

        # 関数を実行
        insert_images_to_excel(output_excel_file, image_dir, regions_and_coords)

    except FileNotFoundError:
        print(f"Error: config file not found at {config_path}")
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from {config_path}. Check file format.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
