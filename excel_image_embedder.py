
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import os
import io

def embed_images_to_excel(excel_file_path, image_folder_path, regions_and_coords):
    """
    指定された画像フォルダの画像をExcelファイルに埋め込みます。
    画像内の特定の領域を抽出し、Excelシートの指定されたセルに貼り付けます。

    Args:
        excel_file_path (str): 生成するExcelファイルのパス。既存のファイルは上書きされます。
        image_folder_path (str): 画像が保存されているフォルダのパス。
        regions_and_coords (list): 領域とセル座標のペアのリスト。
                                   例: [
                                           {"img_region": [x1, y1, x2, y2], "excel_pos": "A1"},
                                           ...
                                       ]
    """
    # 新しいワークブックを作成 (既存ファイルは上書き)
    wb = Workbook()

    # デフォルトで作成されるシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 画像フォルダ内のファイルをリストアップし、ソート
    image_files = [f for f in os.listdir(image_folder_path) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
    image_files.sort() # ファイル名で昇順ソート

    # ソートされたファイルリストを処理
    for filename in image_files:
        image_path = os.path.join(image_folder_path, filename)
        
        try:
            img = Image.open(image_path)
        except Exception as e:
            print(f"Warning: Could not open image {filename}. Skipping. Error: {e}")
            continue

        # 画像ファイル名からシート名を生成 (拡張子を除く)
        sheet_name = os.path.splitext(filename)[0]
        ws = wb.create_sheet(title=sheet_name)

        for item in regions_and_coords:
            img_region = item["img_region"]
            excel_pos = item["excel_pos"]

            try:
                # 画像の領域をクロップ
                cropped_img = img.crop(img_region)

                # openpyxlがファイルパスまたはバイナリデータから画像を読み込むため、
                # クロップした画像を一時的にメモリに保存
                img_byte_arr = io.BytesIO()
                cropped_img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0) # ストリームの先頭に戻る

                # ExcelImageオブジェクトを作成
                excel_img = ExcelImage(img_byte_arr)
                
                # 画像をセルに埋め込む
                ws.add_image(excel_img, excel_pos)
                print(f"Embedded region {img_region} from {filename} to {sheet_name}!{excel_pos}")

            except Exception as e:
                print(f"Warning: Could not crop or embed region {img_region} from {filename} to {excel_pos}. Error: {e}")
                continue

    # Excelファイルを保存
    try:
        wb.save(excel_file_path)
        print(f"Excel file saved successfully to {excel_file_path}")
    except Exception as e:
        print(f"Error: Could not save Excel file to {excel_file_path}. Error: {e}")

import json

if __name__ == '__main__':
    # 設定ファイルを読み込む
    config_path = "/workspace/image_to_office/config.json"
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except FileNotFoundError:
        print(f"Error: Config file not found at {config_path}")
        exit(1)
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from {config_path}")
        exit(1)

    output_excel_path = "/workspace/image_to_office/output_images.xlsx"
    input_image_folder = "/workspace/image_to_office/img"
    
    # 設定から領域とセル座標のペアを取得
    regions_and_coords_from_config = config.get("image_regions_and_coords", [])
    if not regions_and_coords_from_config:
        print("Error: 'image_regions_and_coords' not found in config.json or is empty.")
        exit(1)

    # 関数を実行
    embed_images_to_excel(output_excel_path, input_image_folder, regions_and_coords_from_config)
